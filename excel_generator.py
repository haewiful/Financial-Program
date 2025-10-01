'''
import openpyxl

class ExcelGenerator:
    """Handles the creation and writing of data to an Excel file (.xlsx)."""
    
    def __init__(self, header_list):
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "Data Entry"
        
        # Define headers
        self.headers = header_list
        for col_idx, header in enumerate(self.headers, start=1):
            self.sheet.cell(row=1, column=col_idx, value=header)

    def add_data_row(self, dept, entry, deposit, withdrawal): # TODO update this to dynamic header
        """Adds a new row of data to the Excel sheet, validating types."""
        next_row = self.sheet.max_row + 1
        
        # Validate data types before writing
        try:
            deposit = float(deposit)
            withdrawal = float(withdrawal)
        except ValueError as e:
            # Raise an error that the GUI can catch
            raise ValueError(f"Invalid input: 입금 and 출금 must be a numbers.")
            
        self.sheet.cell(row=next_row, column=1, value=dept)
        self.sheet.cell(row=next_row, column=2, value=entry)
        self.sheet.cell(row=next_row, column=3, value=deposit)
        self.sheet.cell(row=next_row, column=4, value=withdrawal)
        
    def save_file(self, file_path):
        """Saves the workbook to the specified full file path."""
        # Use the provided file_path directly, which is common with filedialog
        try:
            self.workbook.save(file_path)
            return True
        except Exception as e:
            # Print to console for debugging and return False for GUI feedback
            print(f"Error saving file: {e}")
            return False
'''


import openpyxl
import os

class ExcelGenerator:
    """
    Handles the creation, writing, and updating of data in an Excel file (.xlsx).
    The methods are designed to be called by the GUI logic.
    """
    
    def __init__(self, header_list): 
        # Initialize workbook and set headers based on the list passed from main.py
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "Data Entry"
        
        self.headers = header_list 
        
        # Write headers to the first row
        for col_idx, header in enumerate(self.headers, start=1):
            self.sheet.cell(row=1, column=col_idx, value=header)

    def _validate_numeric(self, value, column_name):
        """Helper to validate and coerce numeric data types (handling empty strings as 0)."""
        if value is None or str(value).strip() == "":
            return 0
        try:
            # Using int() for typical accounting/whole dollar values. Use float() if cents are required.
            return int(value) 
        except ValueError:
            raise ValueError(f"'{column_name}' must be a valid whole number.")

    # 🟢 UPDATED: Changed signature to accept dynamic arguments (*data_values)
    def add_data_row(self, *data_values): 
        """Adds a new row of data to the Excel sheet based on positional arguments."""
        next_row = self.sheet.max_row + 1
        
        if len(data_values) != len(self.headers):
            raise ValueError("Data provided does not match the expected number of columns.")

        # Validate and prepare values (assuming numeric columns are the last two: 입금, 출금)
        
        # Validate '입금' (Deposit)
        deposit = self._validate_numeric(data_values[-2], self.headers[-2])
        
        # Validate '출금' (Withdrawal)
        withdrawal = self._validate_numeric(data_values[-1], self.headers[-1])
            
        # Create the final list of values to write: strings first, then validated numbers
        final_values = list(data_values[:-2]) + [deposit, withdrawal]

        # Write data dynamically
        for col_idx, value in enumerate(final_values, start=1):
            self.sheet.cell(row=next_row, column=col_idx, value=value)
            
    # 🟢 NEW: Method required for Treeview editing in main.py
    def update_data_cell(self, user_row_index, col_name, new_value):
        """
        Updates a single cell based on the user-facing row index and column name.
        """
        
        # Find the 1-based column index based on the header name
        try:
            sheet_col = self.headers.index(col_name) + 1
        except ValueError:
            raise ValueError(f"Internal error: Column '{col_name}' not found.")
        
        # Convert user index (1-based, starts at 1) to actual sheet row index (starts at 2)
        sheet_row = user_row_index + 1
        
        # Check if the row index is valid
        if 1 < sheet_row <= self.sheet.max_row:
            
            # Dynamic type validation and coercion
            try:
                if col_name in [self.headers[-2], self.headers[-1]]: # 입금, 출금
                    # Use helper for validation
                    typed_value = self._validate_numeric(new_value, col_name)
                else:
                    # Treat '부서' and '항목' as strings
                    typed_value = str(new_value)
            except ValueError as e:
                raise ValueError(str(e)) # Re-raise error for GUI to display

            # Update the cell in the openpyxl sheet
            self.sheet.cell(row=sheet_row, column=sheet_col, value=typed_value)
            return True
        else:
            raise ValueError(f"Error: Row index {user_row_index} is out of range.")
        
    def save_file(self, file_path):
        """Saves the workbook to the specified full file path."""
        try:
            self.workbook.save(file_path)
            return True
        except Exception as e:
            print(f"Error saving file: {e}")
            return False